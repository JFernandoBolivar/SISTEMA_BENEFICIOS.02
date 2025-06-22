
from flask import Blueprint, render_template, request, redirect, url_for, session, send_file, make_response
import MySQLdb.cursors
from datetime import datetime
from babel.dates import format_date
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import os
from io import BytesIO
from extensions import mysql
from flask import current_app as app

reportes_bp = Blueprint('reportes', __name__)


# listado de entregas

@reportes_bp.route("/listado")
def listado():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute('''
    SELECT d.ID, d.Time_box, d.Staff_ID, d.Observation, d.Lunch,
           personal.Cedula, personal.Code, personal.Name_Com, personal.manually, 
           personal.Location_Admin, personal.Estatus, personal.ESTADOS, personal.Location_Physical, 
           staff.Name_Com AS Registrador_Name,
           autorizados.Nombre AS Nombre_autorizado, autorizados.Cedula AS Cedula_autorizado
    FROM delivery d
    JOIN personal ON d.Data_ID = personal.Cedula 
    LEFT JOIN usuarios ON d.Staff_ID = usuarios.C_I  
    LEFT JOIN personal AS staff ON usuarios.C_I = staff.Cedula  
    LEFT JOIN autorizados ON personal.Cedula = autorizados.beneficiado 
    ''')
    registros = cursor.fetchall()
       
    cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery')
    total_recibido = cursor.fetchone()['total_recibido']
    cursor.close()
    return render_template('tabla.html', registros=registros, total_recibido=total_recibido)

# listado pdf

@reportes_bp.route("/listado_pdf", methods=["GET", "POST"])
def listado_pdf():
    if request.method == "POST":
        fecha = request.form['fecha']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT d.ID, d.Time_box, d.Staff_ID, d.Observation, d.Lunch,
                   personal.Cedula, personal.Name_Com, personal.manually, 
                   personal.Location_Admin, personal.Estatus, personal.ESTADOS, personal.Location_Physical,
                   autorizados.Nombre AS Nombre_autorizado, autorizados.Cedula AS Cedula_autorizado
            FROM delivery d
            JOIN personal ON d.Data_ID = personal.Cedula
            LEFT JOIN autorizados ON personal.ID = autorizados.beneficiado
            WHERE DATE(d.Time_box) = %s
        ''', (fecha,))
        registros = cursor.fetchall()
        cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE DATE(Time_box) = %s', (fecha,))
        total_recibido = cursor.fetchone()['total_recibido']
        cursor.close()

        rendered = render_template('tabla_pdf.html', registros=registros, total_recibido=total_recibido, fecha=fecha)
        pdf = HTML(string=rendered).write_pdf()

        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=listado.pdf'
        return response
    else:
        return render_template('tabla_pdf.html')

# listado de entregados excel

@reportes_bp.route("/listado_excel", methods=["GET", "POST"])
def listado_excel():
    if request.method == "POST":
        fecha = request.form['fecha']
        filtro = request.form.get('filtro', 'todos')

        query = '''
            SELECT d.ID, d.Time_box, d.Staff_ID, d.Observation, d.Lunch,
                   personal.Cedula, personal.Name_Com, personal.manually, 
                   personal.Location_Admin, personal.Estatus, personal.ESTADOS, personal.Location_Physical,
                   autorizados.Nombre AS Nombre_autorizado, autorizados.Cedula AS Cedula_autorizado
            FROM delivery d
            JOIN personal ON d.Data_ID = personal.Cedula
            LEFT JOIN autorizados ON personal.Cedula = autorizados.beneficiado
            WHERE DATE(d.Time_box) = %s
        '''
        if filtro == 'autorizados':
            query += ' AND autorizados.ID IS NOT NULL'
        elif filtro == 'activo':
            query += ' AND personal.Estatus = 1'
        elif filtro == 'pasivo':
            query += ' AND personal.Estatus = 2'
        elif filtro == 'fuera_pais':
            query += ' AND personal.Estatus = 5'
        elif filtro == 'fallecidos':
            query += ' AND personal.Estatus = 6'
        elif filtro == 'requiere_confirmacion':
            query += ' AND personal.Estatus = 7'
        elif filtro == 'suspendidos_tramite':
            query += ' AND personal.Estatus = 3'
        elif filtro == 'comision_vigente':
            query += ' AND personal.Estatus = 10'
        elif filtro == 'comision_vencida':
            query += ' AND personal.Estatus = 9'
        elif filtro == 'manually':
            query += ' AND personal.manually = 1'

        query += ' ORDER BY personal.ESTADOS ASC, personal.Location_Physical ASC, personal.Location_Admin ASC'

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(query, (fecha,))
        registros = cursor.fetchall()
        cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE DATE(Time_box) = %s', (fecha,))
        total_recibido = cursor.fetchone()['total_recibido']
        cursor.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Listado"

        img1_path = os.path.join(app.root_path, 'static/css/img/logo.png')
        img2_path = os.path.join(app.root_path, 'static/css/img/logo2.png')
        img1 = Image(img1_path)
        img2 = Image(img2_path)
        img1.width, img1.height = 60, 60
        img2.width, img2.height = 60, 60
        ws.add_image(img1, 'A1')

        headers = ["#", "Cedula", "Nombre Completo", "Estado", "Estatus", "Unidad Fisica", 
                   "Ubicación administrativa", "Hora de Entrega", "Cedula del Autorizado", 
                   "Nombre del Autorizado", "Observacion", "Registro Manual", "Merienda", 
                   "Cedula del Registrador", "Nombre del Registrador"]
        last_column = chr(64 + len(headers))
        img2.anchor = f'{last_column}1'
        ws.add_image(img2)

        ws.row_dimensions[1].height = 55
        ws.merge_cells(f'A1:{last_column}1')
        ws['A1'] = "Listado de Personas que han Recibido la Caja"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            cell.fill = header_fill

        ws.row_dimensions[2].height = 30

        estatus_map = {
            1: "Activo",
            2: "Pasivo",
            5: "Fuera del país",
            6: "Fallecido",
            7: "Se requiere confirmación",
            3: "Suspendido por trámites administrativos",
            10: "Comisión de Servicio (vigente)",
            9: "Comisión de Servicio (vencida)"
        }

        for idx, registro in enumerate(registros, start=1):
            cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute('SELECT Name_Com FROM personal WHERE Cedula = %s', (registro['Staff_ID'],))
            staff_data = cursor.fetchone()
            cursor.close()

            staff_name = staff_data['Name_Com'] if staff_data else "Desconocido"
            estatus = estatus_map.get(registro['Estatus'], "Desconocido")
           
            row = [
                idx,
                registro['Cedula'],
                registro['Name_Com'],
                registro['ESTADOS'] if registro['Estatus'] in [2, 3, 4, 5, 6, 7, 9, 10] else "",
                estatus,
                registro['Location_Physical'] if registro['Estatus'] == 1 else "",
                registro['Location_Admin'] if registro['Estatus'] == 1 else "",
                registro['Time_box'],
                registro.get('Cedula_autorizado', ''),
                registro.get('Nombre_autorizado', ''),
                registro['Observation'] if registro['Observation'] else ' ',
                "Si" if registro['manually'] else 'No',
                "Si" if registro['Lunch'] else 'No',
                registro['Staff_ID'],
                staff_name
            ]
            
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin'))

        column_widths = {
            'A': 7,
            'B': 10,
            'C': 20,
            'D': 20,
            'E': 25,
            'F': 20,
            'G': 20,
            'H': 20,
            'I': 20,
            'J': 20,
            'K': 20,
            'L': 20,
            'M': 20,
            'N': 20,
            'O': 20
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 82.5

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        nombre_archivo = f"listado_entregados_{fecha_actual}.xlsx"
        
        return send_file(output, download_name=nombre_archivo, as_attachment=True)
    return render_template('tabla_pdf.html')

# listado de las entregas faltantes
@reportes_bp.route("/listado_no_registrado")
def listado_no_registrado():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('''
        SELECT personal.Cedula, personal.Name_Com, personal.Location_Physical, 
               personal.Estatus, personal.ESTADOS, personal.Code, personal.Location_Admin
        FROM personal
        LEFT JOIN delivery ON personal.ID = delivery.Data_ID
        WHERE delivery.ID IS NULL
    ''')
    registros = cursor.fetchall()
    cursor.execute('''
        SELECT COUNT(*) AS total_no_entregados
        FROM personal
        LEFT JOIN delivery ON personal.ID = delivery.Data_ID
        WHERE delivery.ID IS NULL
    ''')
    total_no_entregados = cursor.fetchone()['total_no_entregados']
    cursor.close()
    return render_template('listado_no_registrado.html', registros=registros, total_no_entregados=total_no_entregados)

# listado de las entregas faltantes pdf
@reportes_bp.route("/listado_no_regist_pdf", methods=["GET", "POST"])
def listado_no_regist_pdf():
    filtro = request.args.get('filtro', 'todos')
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    query = '''
        SELECT personal.Cedula, personal.Name_Com, personal.Location_Physical, 
               personal.Location_Admin, personal.Code, personal.Estatus, personal.ESTADOS
        FROM personal
        LEFT JOIN delivery ON personal.Cedula = delivery.Data_ID
        WHERE delivery.ID IS NULL
    '''
    if filtro == 'activos':
        query += ' AND personal.Estatus = 1'
    elif filtro == 'pasivos':
        query += ' AND personal.Estatus = 2'
    elif filtro == 'fuera_pais':
        query += ' AND personal.Estatus = 5'
    elif filtro == 'fallecidos':
        query += ' AND personal.Estatus = 6'
    elif filtro == 'requiere_confirmacion':
        query += ' AND personal.Estatus = 7'
    elif filtro == 'suspendidos_tramite':
        query += ' AND personal.Estatus = 3'
    elif filtro == 'comision_vigente':
        query += ' AND personal.Estatus = 10'
    elif filtro == 'comision_vencida':
        query += ' AND personal.Estatus = 9'

    cursor.execute(query)
    registros = cursor.fetchall()
    
    filtro_count = ''
    if filtro == 'activos':
        filtro_count = 'AND personal.Estatus = 1'
    elif filtro == 'pasivos':
        filtro_count = 'AND personal.Estatus = 2'
    elif filtro == 'fuera_pais':
        filtro_count = 'AND personal.Estatus = 5'
    elif filtro == 'fallecidos':
        filtro_count = 'AND personal.Estatus = 6'
    elif filtro == 'requiere_confirmacion':
        filtro_count = 'AND personal.Estatus = 7'
    elif filtro == 'suspendidos_tramite':
        filtro_count = 'AND personal.Estatus = 3'
    elif filtro == 'comision_vigente':
        filtro_count = 'AND personal.Estatus = 10'
    elif filtro == 'comision_vencida':
        filtro_count = 'AND personal.Estatus = 9'

    cursor.execute(f'''
        SELECT COUNT(*) AS total_no_entregados
        FROM personal
        LEFT JOIN delivery ON personal.ID = delivery.Data_ID
        WHERE delivery.ID IS NULL
        {filtro_count}
    ''')
    total_no_entregados = cursor.fetchone()['total_no_entregados']
    cursor.close()
    
    rendered = render_template('tabla_no_regist_pdf.html', registros=registros, 
                              total_no_entregados=total_no_entregados, filtro=filtro)
    pdf = HTML(string=rendered).write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=listado.pdf'
    return response

# listado de las entregas faltantes excel

@reportes_bp.route("/listado_no_registrado_excel", methods=["GET", "POST"])
def listado_no_registrado_excel():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    
    filtro = request.form.get('filtro', 'todos')
    query = '''
        SELECT personal.Cedula, personal.Name_Com, personal.Location_Physical, 
               personal.Location_Admin, personal.Code, personal.Estatus, personal.ESTADOS
        FROM personal
        LEFT JOIN delivery ON personal.ID = delivery.Data_ID
        WHERE delivery.ID IS NULL
    '''
    if filtro == 'activo':
        query += ' AND personal.Estatus = 1'
    elif filtro == 'pasivo':
        query += ' AND personal.Estatus = 2'
    elif filtro == 'fuera_pais':
        query += ' AND personal.Estatus = 5'
    elif filtro == 'fallecidos':
        query += ' AND personal.Estatus = 6'
    elif filtro == 'requiere_confirmacion':
        query += ' AND personal.Estatus = 7'
    elif filtro == 'suspendidos_tramite':
        query += ' AND personal.Estatus = 3'
    elif filtro == 'comision_vigente':
        query += ' AND personal.Estatus = 10'
    elif filtro == 'comision_vencida':
        query += ' AND personal.Estatus = 9'
    elif filtro == 'autorizados':
        query += ' AND personal.ID IN (SELECT beneficiado FROM autorizados)'

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute(query)
    registros = cursor.fetchall()
    cursor.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "No Entregados"

    img1_path = os.path.join(app.root_path, 'static/css/img/logo.png')
    img2_path = os.path.join(app.root_path, 'static/css/img/logo2.png')
    img1 = Image(img1_path)
    img2 = Image(img2_path)
    img1.width, img1.height = 60, 60
    img2.width, img2.height = 60, 60
    ws.add_image(img1, 'A1')

    headers = ["#", "Cédula", "Nombre Completo", "Unidad Física", "Ubicación Administrativa", 
               "Código", "Estatus", "Estado"]
    last_column = chr(64 + len(headers))
    img2.anchor = f'{last_column}1'
    ws.add_image(img2)

    ws.row_dimensions[1].height = 55
    ws.merge_cells(f'A1:{last_column}1')
    ws['A1'] = "Listado de Personas que No han Recibido la Caja"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        cell.fill = header_fill

    ws.row_dimensions[2].height = 30

    estatus_map = {
        1: "Activo",
        2: "Pasivo",
        3: "Suspendido por trámites administrativos",
        5: "Fuera del país",
        6: "Fallecido",
        7: "Se requiere confirmación",
        9: "Comisión de Servicio (vencida)",
        10: "Comisión de Servicio (vigente)",
        
    }

    for idx, registro in enumerate(registros, start=1):
        estatus = estatus_map.get(registro['Estatus'], "Desconocido")
        row = [
            idx,
            registro['Cedula'],
            registro['Name_Com'],
            registro['Location_Physical'],
            registro['Location_Admin'],
            registro['Code'],
            estatus,
            registro['ESTADOS']
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))

    column_widths = {
        'A': 7,
        'B': 15,
        'C': 25,
        'D': 20,
        'E': 25,
        'F': 15,
        'G': 25,
        'H': 20
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    nombre_archivo = f"listado_no_entregados_{fecha_actual}.xlsx"

    return send_file(output, download_name=nombre_archivo, as_attachment=True)

# reporte de entregas

@reportes_bp.route("/reporte")
def reporte():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))

    mes = request.args.get('mes')
    anio = request.args.get('anio')

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Consulta principal de entregas
    query = '''
        SELECT DATE(Time_box) as fecha,
           SUM(CASE WHEN personal.Estatus = 1 THEN 1 ELSE 0 END) as total_activos,
           SUM(CASE WHEN personal.Estatus = 2 THEN 1 ELSE 0 END) as total_pasivos,
           SUM(CASE WHEN personal.Estatus  IN (9, 11)  THEN 1 ELSE 0 END) as total_comision_vencida,
           SUM(CASE WHEN personal.Estatus = 10 THEN 1 ELSE 0 END) as total_comision_vigente,
           COUNT(*) as total_entregas
    FROM delivery
    JOIN personal ON delivery.Data_ID = personal.Cedula
    '''
    params = []
    where = ""
    if mes and anio:
        where = " WHERE MONTH(Time_box) = %s AND YEAR(Time_box) = %s"
        params.extend([mes, anio])
    query += where + " GROUP BY DATE(Time_box)"

    cursor.execute(query, params)
    reportes = cursor.fetchall()

    # Consulta de entregas de apoyo agrupadas por fecha
    apoyo_query = "SELECT DATE(Fecha) as fecha, SUM(cantidad) as total_apoyo FROM apoyo"
    apoyo_params = []
    if mes and anio:
        apoyo_query += " WHERE MONTH(Fecha) = %s AND YEAR(Fecha) = %s"
        apoyo_params.extend([mes, anio])
    apoyo_query += " GROUP BY DATE(Fecha)"
    cursor.execute(apoyo_query, apoyo_params)
    apoyos = cursor.fetchall()

    apoyo_dict = {str(a['fecha']): a['total_apoyo'] or 0 for a in apoyos}

    # Asegura que todos los reportes tengan total_apoyo y total_entregas_con_apoyo
    for reporte in reportes:
        fecha_str = str(reporte['fecha'])
        reporte['total_apoyo'] = apoyo_dict.get(fecha_str, 0)
        reporte['total_entregas_con_apoyo'] = reporte['total_entregas'] + reporte['total_apoyo']

    # Totales generales
    total_entregas = sum(reporte['total_entregas'] for reporte in reportes)
    total_activos = sum(reporte['total_activos'] for reporte in reportes)
    total_pasivos = sum(reporte['total_pasivos'] for reporte in reportes)
    total_comision_vencida = sum(reporte['total_comision_vencida'] for reporte in reportes)
    total_comision_vigente = sum(reporte['total_comision_vigente'] for reporte in reportes)
    total_apoyo = sum(reporte.get('total_apoyo', 0) for reporte in reportes)
    total_entregas_con_apoyo = sum(reporte['total_entregas_con_apoyo'] for reporte in reportes)
    cursor.close()

    for reporte in reportes:
        fecha = reporte['fecha']
        reporte['fecha_formateada'] = format_date(fecha, format='full', locale='es_ES')

    return render_template(
        'reporte.html',
        reportes=reportes,
        total_entregas=total_entregas,
        total_activos=total_activos,
        total_pasivos=total_pasivos,
        total_comision_vencida=total_comision_vencida,
        total_comision_vigente=total_comision_vigente,
        total_apoyo=total_apoyo,
        total_entregas_con_apoyo=total_entregas_con_apoyo,
        mes=mes,
        anio=anio
    )

# pdf para los reportes


@reportes_bp.route("/reporte_pdf", methods=["GET", "POST"])
def reporte_pdf():
    # Puedes obtener mes y año por GET o POST según tu formulario
    mes = request.values.get('mes')
    anio = request.values.get('anio')

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Consulta principal de entregas
    query = '''
        SELECT DATE(Time_box) as fecha,
               SUM(CASE WHEN personal.Estatus = 1 THEN 1 ELSE 0 END) as total_activos,
               SUM(CASE WHEN personal.Estatus = 2 THEN 1 ELSE 0 END) as total_pasivos,
               SUM(CASE WHEN personal.Estatus = 10 THEN 1 ELSE 0 END) as total_comision_vigente,
               SUM(CASE WHEN personal.Estatus = 9 THEN 1 ELSE 0 END) as total_comision_vencida,
               COUNT(*) as total_entregas
        FROM delivery
        JOIN personal ON delivery.Data_ID = personal.Cedula
    '''
    params = []
    if mes and anio:
        query += " WHERE MONTH(Time_box) = %s AND YEAR(Time_box) = %s"
        params.extend([mes, anio])
    query += " GROUP BY DATE(Time_box)"

    cursor.execute(query, params)
    reportes = cursor.fetchall()

    # Consulta de entregas de apoyo agrupadas por fecha
    apoyo_query = "SELECT DATE(Fecha) as fecha, SUM(cantidad) as total_apoyo FROM apoyo"
    apoyo_params = []
    if mes and anio:
        apoyo_query += " WHERE MONTH(Fecha) = %s AND YEAR(Fecha) = %s"
        apoyo_params.extend([mes, anio])
    apoyo_query += " GROUP BY DATE(Fecha)"
    cursor.execute(apoyo_query, apoyo_params)
    apoyos = cursor.fetchall()

    apoyo_dict = {str(a['fecha']): a['total_apoyo'] or 0 for a in apoyos}

    # Asegura que todos los reportes tengan total_apoyo y total_entregas_con_apoyo
    for reporte in reportes:
        fecha_str = str(reporte['fecha'])
        reporte['total_apoyo'] = apoyo_dict.get(fecha_str, 0)
        reporte['total_entregas_con_apoyo'] = reporte['total_entregas'] + reporte['total_apoyo']

    # Totales generales
    total_entregas = sum(r['total_entregas'] for r in reportes)
    total_activos = sum(r['total_activos'] for r in reportes)
    total_pasivos = sum(r['total_pasivos'] for r in reportes)
    total_comision_vigente = sum(r['total_comision_vigente'] for r in reportes)
    total_comision_vencida = sum(r['total_comision_vencida'] for r in reportes)
    total_apoyo = sum(r.get('total_apoyo', 0) for r in reportes)
    total_entregas_con_apoyo = sum(r['total_entregas_con_apoyo'] for r in reportes)

    cursor.close()

    for reporte in reportes:
        fecha = reporte['fecha']
        reporte['fecha_formateada'] = format_date(fecha, format='full', locale='es_ES')

    rendered = render_template(
        'reporte_pdf.html',
        reportes=reportes,
        total_entregas=total_entregas,
        total_activos=total_activos,
        total_pasivos=total_pasivos,
        total_comision_vigente=total_comision_vigente,
        total_comision_vencida=total_comision_vencida,
        total_apoyo=total_apoyo,
        total_entregas_con_apoyo=total_entregas_con_apoyo,
        mes=mes,
        anio=anio
    )
    pdf = HTML(string=rendered).write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=reporte.pdf'
    return response
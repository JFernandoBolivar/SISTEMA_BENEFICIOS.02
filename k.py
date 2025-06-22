# reportes/routes.py
from flask import Blueprint, render_template, request, redirect, url_for, session, send_file, make_response
import MySQLdb.cursors
from datetime import datetime
from babel.dates import format_date
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import os
from io import BytesIO
from extensions import mysql
from flask import current_app as app

reportes_bp = Blueprint('reportes', __name__)

# CAMBIOS PRINCIPALES:
# 1. Todas las referencias a personal.ID cambiadas a personal.Cedula
# 2. Join con usuarios para obtener datos del registrador
# 3. Cambio de ESTADOS a ESTADO
# 4. Ajuste en las relaciones con autorizados

@reportes_bp.route("/reporte")
def reporte():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute('''
    SELECT DATE(Time_box) as fecha,
           SUM(CASE WHEN personal.Estatus = 1 THEN 1 ELSE 0 END) as total_activos,
           SUM(CASE WHEN personal.Estatus = 2 THEN 1 ELSE 0 END) as total_pasivos,
           SUM(CASE WHEN personal.Estatus IN (3,4) THEN 1 ELSE 0 END) as total_comision_vencida,
           SUM(CASE WHEN personal.Estatus = 5 THEN 1 ELSE 0 END) as total_comision_vigente,
           COUNT(*) as total_entregas
    FROM delivery
    JOIN personal ON delivery.Data_ID = personal.Cedula  -- Cambiado a Cedula
    GROUP BY DATE(Time_box)
    ''')
    reportes = cursor.fetchall()

    total_entregas = sum(reporte['total_entregas'] for reporte in reportes)
    total_activos = sum(reporte['total_activos'] for reporte in reportes)
    total_pasivos = sum(reporte['total_pasivos'] for reporte in reportes)
    total_comision_vencida = sum(reporte['total_comision_vencida'] for reporte in reportes)
    total_comision_vigente = sum(reporte['total_comision_vigente'] for reporte in reportes)

    cursor.close()

    for reporte in reportes:
        fecha = reporte['fecha']
        reporte['fecha_formateada'] = format_date(fecha, format='full', locale='es_ES')

    return render_template(
        'reporte.html',
        reportes=reportes,
        total_entregas=total_entregas,
        total_activos=total_activos,
        total_pasivos=total_pasivos,
        total_comision_vencida=total_comision_vencida,
        total_comision_vigente=total_comision_vigente
    )

@reportes_bp.route("/reporte_pdf", methods=["GET", "POST"])
def reporte_pdf():
    if request.method == "POST":
        fecha = request.form['fecha']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT DATE(Time_box) as fecha,
                   SUM(CASE WHEN personal.Estatus = 1 THEN 1 ELSE 0 END) as total_activos,
                   SUM(CASE WHEN personal.Estatus = 2 THEN 1 ELSE 0 END) as total_pasivos,
                   SUM(CASE WHEN personal.Estatus IN (3,4) THEN 1 ELSE 0 END) as total_comision_vencida,
                   SUM(CASE WHEN personal.Estatus = 5 THEN 1 ELSE 0 END) as total_comision_vigente,
                   COUNT(*) as total_entregas
            FROM delivery
            JOIN personal ON delivery.Data_ID = personal.Cedula  -- Cambiado a Cedula
            WHERE DATE(Time_box) = %s
            GROUP BY DATE(Time_box)
        ''', (fecha,))
    else:
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT DATE(Time_box) as fecha,
                   SUM(CASE WHEN personal.Estatus = 1 THEN 1 ELSE 0 END) as total_activos,
                   SUM(CASE WHEN personal.Estatus = 2 THEN 1 ELSE 0 END) as total_pasivos,
                   SUM(CASE WHEN personal.Estatus IN (3,4) THEN 1 ELSE 0 END) as total_comision_vencida,
                   SUM(CASE WHEN personal.Estatus = 5 THEN 1 ELSE 0 END) as total_comision_vigente,
                   COUNT(*) as total_entregas
            FROM delivery
            JOIN personal ON delivery.Data_ID = personal.Cedula  -- Cambiado a Cedula
            GROUP BY DATE(Time_box)
        ''')
    reportes = cursor.fetchall()
    cursor.close()
    
    total_entregas = sum(reporte['total_entregas'] for reporte in reportes)
    total_activos = sum(reporte['total_activos'] for reporte in reportes)
    total_pasivos = sum(reporte['total_pasivos'] for reporte in reportes)
    total_comision_vencida = sum(reporte['total_comision_vencida'] for reporte in reportes)
    total_comision_vigente = sum(reporte['total_comision_vigente'] for reporte in reportes)
    
    for reporte in reportes:
        fecha = reporte['fecha']
        reporte['fecha_formateada'] = format_date(fecha, format='full', locale='es_ES')
    
    rendered = render_template(
        'reporte_pdf.html',
        reportes=reportes,
        total_entregas=total_entregas,
        total_activos=total_activos,
        total_pasivos=total_pasivos,
        total_comision_vencida=total_comision_vencida,
        total_comision_vigente=total_comision_vigente
    )
    pdf = HTML(string=rendered).write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=reporte.pdf'
    return response

@reportes_bp.route("/listado")
def listado():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cursor.execute('''
    SELECT d.ID, d.Time_box, d.Staff_ID, d.Observation, d.Lunch,
           personal.Cedula, personal.Code, personal.Name_Com, personal.manually, 
           personal.Location_Admin, personal.Estatus, personal.ESTADO, personal.Location_Physical, 
           staff.Name_Com AS Registrador_Name,
           autorizados.Nombre AS Nombre_autorizado, autorizados.Cedula AS Cedula_autorizado
    FROM delivery d
    JOIN personal ON d.Data_ID = personal.Cedula  -- Cambiado a Cedula
    LEFT JOIN usuarios ON d.Staff_ID = usuarios.C_I  -- Nueva relación con usuarios
    LEFT JOIN personal AS staff ON usuarios.C_I = staff.Cedula  -- Obtener datos del registrador
    LEFT JOIN autorizados ON personal.Cedula = autorizados.beneficiado  -- Cambiado a Cedula
    ''')
    registros = cursor.fetchall()
       
    cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery')
    total_recibido = cursor.fetchone()['total_recibido']
    cursor.close()
    return render_template('tabla.html', registros=registros, total_recibido=total_recibido)

@reportes_bp.route("/listado_pdf", methods=["GET", "POST"])
def listado_pdf():
    if request.method == "POST":
        fecha = request.form['fecha']
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT d.ID, d.Time_box, d.Staff_ID, d.Observation, d.Lunch,
                   personal.Cedula, personal.Name_Com, personal.manually, 
                   personal.Location_Admin, personal.Estatus, personal.ESTADO, personal.Location_Physical,  -- 
                   autorizados.Nombre AS Nombre_autorizado, autorizados.Cedula AS Cedula_autorizado
            FROM delivery d
            JOIN personal ON d.Data_ID = personal.Cedula  -- Cambiado a Cedula
            LEFT JOIN autorizados ON personal.Cedula = autorizados.beneficiado  -- Cambiado a Cedula
            WHERE DATE(d.Time_box) = %s
        ''', (fecha,))
        registros = cursor.fetchall()
        cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE DATE(Time_box) = %s', (fecha,))
        total_recibido = cursor.fetchone()['total_recibido']
        cursor.close()

        rendered = render_template('tabla_pdf.html', registros=registros, total_recibido=total_recibido, fecha=fecha)
        pdf = HTML(string=rendered).write_pdf()

        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'inline; filename=listado.pdf'
        return response
    else:
        return render_template('tabla_pdf.html')
    
@reportes_bp.route("/listado_excel", methods=["GET", "POST"])
def listado_excel():
    if request.method == "POST":
        fecha = request.form['fecha']
        filtro = request.form.get('filtro', 'todos')

        query = '''
            SELECT d.ID, d.Time_box, d.Staff_ID, d.Observation, d.Lunch,
                   personal.Cedula, personal.Name_Com, personal.manually, 
                   personal.Location_Admin, personal.Estatus, personal.ESTADO, personal.Location_Physical,  -- Cambiado ESTADOS -> ESTADO
                   autorizados.Nombre AS Nombre_autorizado, autorizados.Cedula AS Cedula_autorizado
            FROM delivery d
            JOIN personal ON d.Data_ID = personal.Cedula  -- Cambiado a Cedula
            LEFT JOIN autorizados ON personal.Cedula = autorizados.beneficiado  -- Cambiado a Cedula
            WHERE DATE(d.Time_box) = %s
        '''
        if filtro == 'autorizados':
            query += ' AND autorizados.ID IS NOT NULL'
        elif filtro == 'activo':
            query += ' AND personal.Estatus = 1'
        elif filtro == 'pasivo':
            query += ' AND personal.Estatus = 2'
        elif filtro == 'manually':
            query += ' AND personal.manually = 1'
        elif filtro == 'comision_vencida':
            query += ' AND personal.Estatus IN (3,4)'
        elif filtro == 'comision_vigente':
            query += ' AND personal.Estatus = 5'

        query += ' ORDER BY personal.ESTADO ASC, personal.Location_Physical ASC, personal.Location_Admin ASC'  
        

        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(query, (fecha,))
        registros = cursor.fetchall()
        cursor.execute('SELECT COUNT(*) AS total_recibido FROM delivery WHERE DATE(Time_box) = %s', (fecha,))
        total_recibido = cursor.fetchone()['total_recibido']
        cursor.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Listado"

        img1_path = os.path.join(app.root_path, 'static/css/img/logo.png')
        img2_path = os.path.join(app.root_path, 'static/css/img/logo2.png')
        img1 = Image(img1_path)
        img2 = Image(img2_path)
        img1.width, img1.height = 60, 60
        img2.width, img2.height = 60, 60
        ws.add_image(img1, 'A1')

        headers = ["#", "Cedula", "Nombre Completo", "Estado", "Estatus", "Unidad Fisica", 
                   "Ubicación administrativa", "Hora de Entrega", "Cedula del Autorizado", 
                   "Nombre del Autorizado", "Observacion", "Registro Manual", "Merienda", 
                   "Cedula del Registrador", "Nombre del Registrador"]
        last_column = chr(64 + len(headers))
        img2.anchor = f'{last_column}1'
        ws.add_image(img2)

        ws.row_dimensions[1].height = 55
        ws.merge_cells(f'A1:{last_column}1')
        ws['A1'] = "Listado de Personas que han Recibido la Caja"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))
            cell.fill = header_fill

        ws.row_dimensions[2].height = 30

        for idx, registro in enumerate(registros, start=1):
            # Obtener nombre del registrador mediante usuarios
            cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute('''
                SELECT p.Name_Com 
                FROM usuarios u
                JOIN personal p ON u.C_I = p.Cedula
                WHERE u.C_I = %s
            ''', (registro['Staff_ID'],))
            staff_data = cursor.fetchone()
            cursor.close()

            staff_name = staff_data['Name_Com'] if staff_data else "Desconocido"
            
            estatus_map = {
                1: "Activo",
                2: "Pasivo",
                3: "Comisión Vencida",
                4: "Comisión Vencida",
                5: "Comisión Vigente"
            }
            estatus = estatus_map.get(registro['Estatus'], "Desconocido")
           
            row = [
                idx,
                registro['Cedula'],
                registro['Name_Com'],
                registro['ESTADO'] if registro['Estatus'] in [2, 3, 4, 5] else "",  # Cambiado ESTADOS -> ESTADO
                estatus,
                registro['Location_Physical'] if registro['Estatus'] == 1 else "",
                registro['Location_Admin'] if registro['Estatus'] == 1 else "",
                registro['Time_box'],
                registro.get('Cedula_autorizado', ''),
                registro.get('Nombre_autorizado', ''),
                registro['Observation'] if registro['Observation'] else ' ',
                "Si" if registro['manually'] else 'No',
                "Si" if registro['Lunch'] else 'No',
                registro['Staff_ID'],
                staff_name
            ]
            
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin'))

        column_widths = {
            'A': 7,
            'B': 10,
            'C': 20,
            'D': 20,
            'E': 20,
            'F': 20,
            'G': 20,
            'H': 20,
            'I': 20,
            'J': 20,
            'K': 20,
            'L': 20,
            'M': 20,
            'N': 20,
            'O': 20
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 82.5

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fecha_actual = datetime.now().strftime('%Y-%m-%d')
        nombre_archivo = f"listado_entregados_{fecha_actual}.xlsx"
        
        return send_file(output, download_name=nombre_archivo, as_attachment=True)
    return render_template('tabla_pdf.html')

@reportes_bp.route("/listado_no_registrado")
def listado_no_registrado():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute('''
        SELECT personal.Cedula, personal.Name_Com, personal.Location_Physical, 
               personal.Estatus, personal.ESTADO, personal.Code, personal.Location_Admin  -- Cambiado ESTADOS -> ESTADO
        FROM personal
        LEFT JOIN delivery ON personal.Cedula = delivery.Data_ID  -- Cambiado a Cedula
        WHERE delivery.ID IS NULL
    ''')
    registros = cursor.fetchall()
    cursor.execute('''
        SELECT COUNT(*) AS total_no_entregados
        FROM personal
        LEFT JOIN delivery ON personal.Cedula = delivery.Data_ID  -- Cambiado a Cedula
        WHERE delivery.ID IS NULL
    ''')
    total_no_entregados = cursor.fetchone()['total_no_entregados']
    cursor.close()
    return render_template('listado_no_registrado.html', registros=registros, total_no_entregados=total_no_entregados)

@reportes_bp.route("/listado_no_regist_pdf", methods=["GET", "POST"])
def listado_no_regist_pdf():
    filtro = request.args.get('filtro', 'todos')
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    query = '''
        SELECT personal.Cedula, personal.Name_Com, personal.Location_Physical, 
               personal.Location_Admin, personal.Code, personal.Estatus, personal.ESTADO  -- Cambiado ESTADOS -> ESTADO
        FROM personal
        LEFT JOIN delivery ON personal.Cedula = delivery.Data_ID  -- Cambiado a Cedula
        WHERE delivery.ID IS NULL
    '''
    if filtro == 'activos':
        query += ' AND personal.Estatus = 1'
    elif filtro == 'pasivos':
        query += ' AND personal.Estatus = 2'
    elif filtro == 'comision_vigente':
        query += ' AND personal.Estatus = 5'
    elif filtro == 'comision_vencida':
        query += ' AND personal.Estatus IN (3,4)'
    
    cursor.execute(query)
    registros = cursor.fetchall()
    
    cursor.execute(f'''
        SELECT COUNT(*) AS total_no_entregados
        FROM personal
        LEFT JOIN delivery ON personal.Cedula = delivery.Data_ID  -- Cambiado a Cedula
        WHERE delivery.ID IS NULL
        {'AND personal.Estatus = 1' if filtro == 'activos' else ''}
        {'AND personal.Estatus = 2' if filtro == 'pasivos' else ''}
        {'AND personal.Estatus = 5' if filtro == 'comision_vigente' else ''}
        {'AND personal.Estatus IN (3,4)' if filtro == 'comision_vencida' else ''}
    ''')
    total_no_entregados = cursor.fetchone()['total_no_entregados']
    cursor.close()
    
    rendered = render_template('tabla_no_regist_pdf.html', registros=registros, 
                              total_no_entregados=total_no_entregados, filtro=filtro)
    pdf = HTML(string=rendered).write_pdf()

    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'inline; filename=listado.pdf'
    return response

@reportes_bp.route("/listado_no_registrado_excel", methods=["GET", "POST"])
def listado_no_registrado_excel():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    
    filtro = request.form.get('filtro', 'todos')
    query = '''
        SELECT personal.Cedula, personal.Name_Com, personal.Location_Physical, 
               personal.Location_Admin, personal.Code, personal.Estatus, personal.ESTADO  -- Cambiado ESTADOS -> ESTADO
        FROM personal
        LEFT JOIN delivery ON personal.Cedula = delivery.Data_ID  -- Cambiado a Cedula
        WHERE delivery.ID IS NULL
    '''
    if filtro == 'activo':
        query += ' AND personal.Estatus = 1'
    elif filtro == 'pasivo':
        query += ' AND personal.Estatus = 2'
    elif filtro == 'comision_vigente':
        query += ' AND personal.Estatus = 5'
    elif filtro == 'comision_vencida':
        query += ' AND personal.Estatus IN (3,4)'
    elif filtro == 'autorizados':
        query += ' AND personal.Cedula IN (SELECT beneficiado FROM autorizados)'  # Cambiado a Cedula

    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cursor.execute(query)
    registros = cursor.fetchall()
    cursor.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "No Entregados"

    img1_path = os.path.join(app.root_path, 'static/css/img/logo.png')
    img2_path = os.path.join(app.root_path, 'static/css/img/logo2.png')
    img1 = Image(img1_path)
    img2 = Image(img2_path)
    img1.width, img1.height = 60, 60
    img2.width, img2.height = 60, 60
    ws.add_image(img1, 'A1')

    headers = ["#", "Cédula", "Nombre Completo", "Unidad Física", "Ubicación Administrativa", 
               "Código", "Estatus", "Estado"]
    last_column = chr(64 + len(headers))
    img2.anchor = f'{last_column}1'
    ws.add_image(img2)

    ws.row_dimensions[1].height = 55
    ws.merge_cells(f'A1:{last_column}1')
    ws['A1'] = "Listado de Personas que No han Recibido la Caja"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws.append(headers)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        cell.fill = header_fill

    ws.row_dimensions[2].height = 30

    estatus_map = {
        1: "Activo",
        2: "Pasivo",
        3: "Comisión Vencida",
        4: "Comisión Vencida",
        5: "Comisión Vigente"
    }

    for idx, registro in enumerate(registros, start=1):
        estatus = estatus_map.get(registro['Estatus'], "Desconocido")
        row = [
            idx,
            registro['Cedula'],
            registro['Name_Com'],
            registro['Location_Physical'],
            registro['Location_Admin'],
            registro['Code'],
            estatus,
            registro['ESTADO']  # Cambiado ESTADOS -> ESTADO
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))

    column_widths = {
        'A': 7,
        'B': 15,
        'C': 25,
        'D': 20,
        'E': 25,
        'F': 15,
        'G': 20,
        'H': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    nombre_archivo = f"listado_no_entregados_{fecha_actual}.xlsx"

    return send_file(output, download_name=nombre_archivo, as_attachment=True)
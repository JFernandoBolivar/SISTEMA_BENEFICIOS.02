from flask import Blueprint, render_template, request, redirect, url_for, session, jsonify, flash, send_file
import MySQLdb.cursors
from datetime import datetime
from babel.dates import format_date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from extensions import mysql

gestion_autorizados_bp = Blueprint('gestion_autorizados', __name__)

@gestion_autorizados_bp.route("/listado_autorizados", methods=["GET"])
def listado_autorizados():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Consulta adaptada al nuevo esquema
    cursor.execute('''
        SELECT 
            p.Cedula AS Beneficiario_Cedula,
            p.Name_Com AS Beneficiario_Nombre,
            a.Cedula AS Autorizado_Cedula,
            a.Nombre AS Autorizado_Nombre
        FROM autorizados a
        JOIN personal p ON a.beneficiado = p.Cedula
    ''')
    
    registros = cursor.fetchall()
    cursor.close()
    
    total_autorizados = len(registros)
    return render_template('listado_autorizados.html', 
                          registros=registros,
                          total_autorizados=total_autorizados)

@gestion_autorizados_bp.route("/editar_autorizado/<int:beneficiado_cedula>", methods=["GET", "POST"])
def editar_autorizado(beneficiado_cedula):
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if request.method == "POST":
        nuevo_nombre = request.form['nuevo_nombre'].strip()
        nueva_cedula = request.form['nueva_cedula'].strip()
        
        # Actualizar el autorizado
        cursor.execute('''
            UPDATE autorizados
            SET Nombre = %s, Cedula = %s
            WHERE beneficiado = %s
        ''', (nuevo_nombre, nueva_cedula, beneficiado_cedula))
        
        mysql.connection.commit()
        cursor.close()
        return redirect(url_for('gestion_autorizados.listado_autorizados'))
    
    # Obtener el beneficiario y su autorizado
    cursor.execute('''
        SELECT 
            p.Cedula AS Beneficiario_Cedula,
            p.Name_Com AS Beneficiario_Nombre,
            a.Cedula AS Autorizado_Cedula,
            a.Nombre AS Autorizado_Nombre
        FROM personal p
        JOIN autorizados a ON p.Cedula = a.beneficiado
        WHERE a.beneficiado = %s
    ''', (beneficiado_cedula,))
    
    registro = cursor.fetchone()
    cursor.close()
    
    if not registro:
        flash("No se encontró el autorizado", "danger")
        return redirect(url_for('gestion_autorizados.listado_autorizados'))
    
    return render_template('editar_autorizado.html', registro=registro)

@gestion_autorizados_bp.route("/reporte_entregas_usuario", methods=["GET", "POST"])
def reporte_entregas_usuario():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    fecha = request.form.get('fecha', None)
    
    if fecha:
      
        cursor.execute('''
            SELECT 
                u.C_I AS Staff_ID, 
                DATE(d.Time_box) as fecha, 
                COUNT(*) as total_entregas,
                p.Name_Com AS staff_name
            FROM delivery d
            JOIN usuarios u ON d.Staff_ID = u.C_I
            JOIN personal p ON u.C_I = p.Cedula
            WHERE DATE(d.Time_box) = %s
            GROUP BY u.C_I, DATE(d.Time_box)
        ''', (fecha,))
    else:
        cursor.execute('''
            SELECT 
                u.C_I AS Staff_ID, 
                DATE(d.Time_box) as fecha, 
                COUNT(*) as total_entregas,
                p.Name_Com AS staff_name
            FROM delivery d
            JOIN usuarios u ON d.Staff_ID = u.C_I
            JOIN personal p ON u.C_I = p.Cedula
            GROUP BY u.C_I, DATE(d.Time_box)
        ''')
    
    reportes = cursor.fetchall()
    cursor.close()
    
    # Formatear fechas
    for reporte in reportes:
        fecha = reporte['fecha']
        reporte['fecha_formateada'] = format_date(fecha, format='full', locale='es_ES')
    
    return render_template('reporte_entregas_usuario.html', reportes=reportes)

@gestion_autorizados_bp.route("/reporte_entregas_usuario_excel", methods=["GET", "POST"])
def reporte_entregas_usuario_excel():
    if 'loggedin' not in session:
        return redirect(url_for('auth.login'))
    
    fecha = request.form.get('fecha', None)
    
    cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    if fecha:
        cursor.execute('''
            SELECT 
                u.C_I AS Staff_ID, 
                DATE(d.Time_box) as fecha, 
                COUNT(*) as total_entregas,
                p.Name_Com AS staff_name
            FROM delivery d
            JOIN usuarios u ON d.Staff_ID = u.C_I
            JOIN personal p ON u.C_I = p.Cedula
            WHERE DATE(d.Time_box) = %s
            GROUP BY u.C_I, DATE(d.Time_box)
        ''', (fecha,))
    else:
        cursor.execute('''
            SELECT 
                u.C_I AS Staff_ID, 
                DATE(d.Time_box) as fecha, 
                COUNT(*) as total_entregas,
                p.Name_Com AS staff_name
            FROM delivery d
            JOIN usuarios u ON d.Staff_ID = u.C_I
            JOIN personal p ON u.C_I = p.Cedula
            GROUP BY u.C_I, DATE(d.Time_box)
        ''')
    
    reportes = cursor.fetchall()
    cursor.close()
    
    # Formatear fechas manualmente
    dias = {
        'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
        'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado',
        'Sunday': 'Domingo'
    }
    meses = {
        'January': 'enero', 'February': 'febrero', 'March': 'marzo',
        'April': 'abril', 'May': 'mayo', 'June': 'junio',
        'July': 'julio', 'August': 'agosto', 'September': 'septiembre',
        'October': 'octubre', 'November': 'noviembre', 'December': 'diciembre'
    }
    
    for reporte in reportes:
        fecha = reporte['fecha']
        fecha_str = fecha.strftime('%A, %d de %B de %Y')
        
        for eng, esp in dias.items():
            fecha_str = fecha_str.replace(eng, esp)
        for eng, esp in meses.items():
            fecha_str = fecha_str.replace(eng, esp)
        
        reporte['fecha_formateada'] = fecha_str
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Entregas Usuario"

    headers = ["Fecha", "Usuario", "Total Entregas"]
    ws.append(headers)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        cell.fill = header_fill

    for reporte in reportes:
        row = [
            reporte['fecha_formateada'],
            reporte['staff_name'],
            reporte['total_entregas']
        ]
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    column_widths = {
        'A': 30,  # Más ancho para la fecha formateada
        'B': 40,  # Más ancho para el nombre
        'C': 20   # Ancho para el total
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_actual = datetime.now().strftime('%Y-%m-%d')
    nombre_archivo = f"Reporte_entregas_por_usuario_{fecha_actual}.xlsx"

    return send_file(output, download_name=nombre_archivo, as_attachment=True)